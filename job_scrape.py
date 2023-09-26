import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import csv
import pandas as pd
from selenium.common.exceptions import NoSuchElementException
from typing import cast

from selenium.webdriver.support.ui import WebDriverWait
import re
import openpyxl as xl


def get_version():
  """
  Gets the version of the Selenium library that is installed.pip install selenium

  Returns:
    The version of the Selenium library.
  """
  return webdriver.__version__

def login_to_linkedin(driver, email, password):
  """
  Logs in to LinkedIn.

  Args:
    driver: The Selenium webdriver.
    email: The user's email address.
    password: The user's password.
  """

  driver.get("https://www.linkedin.com/home")
  time.sleep(5)

  user_element = driver.find_element("id", "session_key")
  user_element.send_keys(email)

  password_element = driver.find_element("id", "session_password")
  password_element.send_keys(password)

  password_element.submit()

def search_for_jobs(driver):
  """
  Searches for jobs on LinkedIn.

  Args:
    driver: The Selenium webdriver.
    keywords: A list of keywords.
  """

  driver.get("https://www.linkedin.com/jobs/search/?currentJobId=3716300385&f_TPR=r2592000&geoId=90009496&keywords=data%20scientist&location=London%20Area%2C%20United%20Kingdom&origin=JOB_SEARCH_PAGE_KEYWORD_AUTOCOMPLETE&refresh=true")
  time.sleep(3)

def get_job_ids(driver):
  """
  gets all the ids for jobs listing on the page.

  Args:
    driver: The Selenium webdriver.
    keywords: A list of keywords.

  returns a list of job ids
  """
  card_container = driver.find_element(By.CLASS_NAME, "scaffold-layout__list-container")
  job_cards = card_container.find_elements(By.TAG_NAME, "li")

  job_ids = []
  for job_card in job_cards:
    job_id = job_card.get_attribute("id")
    if job_id != '':
      job_ids.append(job_id)

  return job_ids

def check_words(job_description,keywords):
  """
  returns a list of 1s and 0s corresponing to if a the word in keywords is present in the job desctiprion
     
  Args:
    driver: The Selenium webdriver.
    keywords: A list of keywords.
  """
  presence_indicator = []

  for word in keywords:
    if word in job_description:
      presence_indicator.append(1)
    else:
      presence_indicator.append(0)

  return presence_indicator

def go_to_next_page(driver):
  """
  Goes to the next page of job results.

  Args:
    driver: The Selenium webdriver.
  """

  pagination_element = driver.find_element(By.CLASS_NAME, "jobs-search-results-list__pagination")
  buttons = pagination_element.find_elements(By.TAG_NAME, "button")
 
  # Find the button that represents the current page.
  current_button = None
  for button in buttons:
    if button.get_attribute("aria-current") == "true":
      current_button = button
      break

   # Check if the current button is at the end of the list.
  if current_button is None or buttons.index(current_button) == len(buttons) - 1:
    return False

  # Get the next button.
  next_button = buttons[buttons.index(current_button) + 1]
  # Click on the next button.
  next_button.click()
  return True
def class_exists(driver, class_name):
  """
  Checks if a class exists on the current page.

  Args:
    driver: The Selenium webdriver.
    class_name: The name of the class to check.

  Returns:
    True if the class exists, False otherwise.
  """

  try:
    driver.find_element(By.CLASS_NAME, class_name)
    return True
  except NoSuchElementException:
    return False


def split_job_post(string):
  """Splits a job post string into company name, location, date of post, and number of applicants.

  Args:
    string: A string containing the job post.

  Returns:
    A tuple containing the company name, location, date of post, and number of applicants.
  """

  # Split the string into a list of words.
  parts = string.split('·')

  # The company name is the first word.
  company_name = parts[0]

  location_date = parts[1]
  words= location_date.split()
  index = next((i for i, word in enumerate(words) if word == 'reposted' or re.search(r'\d', word)), None)
  location = ' '.join(words[:index])
  time_of_post = ' '.join(words[index:])

  number_of_applicants = parts[2]
  return company_name, location, time_of_post, number_of_applicants

def save_to_excel(job_data, filename,keywords):
  """Saves job data to an Excel spreadsheet.

  Args:
    job_data: A list of job data, where each job data is a list of the following:
      * Job title
      * Company name
      * Location
      * Time of post
      * Number of applicants
      * Python keyword presence
      * Machine learning keyword presence
      * Data science keyword presence
    filename: The name of the Excel spreadsheet to save the data to.
  """

  wb = xl.Workbook()
  ws = wb.active

  # Write the column headers.
  ws.cell(row=1, column=1).value = "Job Title"
  ws.cell(row=1, column=2).value = "Company Name"
  ws.cell(row=1, column=3).value = "Location"
  ws.cell(row=1, column=4).value = "Time of Post"
  ws.cell(row=1, column=5).value = "Number of Applicants"
  ws.cell(row=1, column=6).value = keywords[0]
  ws.cell(row=1, column=7).value = keywords[1]
  ws.cell(row=1, column=8).value = keywords[2]
  ws.cell(row=1, column=9).value = keywords[3]
  ws.cell(row=1, column=10).value = keywords[4]
  ws.cell(row=1, column=11).value = keywords[5]
  ws.cell(row=1, column=12).value = keywords[6]
  ws.cell(row=1, column=13).value = keywords[7]
  ws.cell(row=1, column=14).value = keywords[8]
  ws.cell(row=1, column=15).value = keywords[9]
  ws.cell(row=1, column=16).value = keywords[10]
  ws.cell(row=1, column=17).value = keywords[11]
  ws.cell(row=1, column=18).value = keywords[12]


  # Write the job data to the spreadsheet.
  for i in range(len(job_data)):
    job_datum = job_data[i]
    ws.cell(row=i + 2, column=1).value = job_datum[0]
    ws.cell(row=i + 2, column=2).value = job_datum[1]
    ws.cell(row=i + 2, column=3).value = job_datum[2]
    ws.cell(row=i + 2, column=4).value = job_datum[3]
    ws.cell(row=i + 2, column=5).value = job_datum[4]
    ws.cell(row=i + 2, column=6).value = job_datum[5]
    ws.cell(row=i + 2, column=7).value = job_datum[6]
    ws.cell(row=i + 2, column=8).value = job_datum[7]
    ws.cell(row=i + 2, column=9).value = job_datum[8]
    ws.cell(row=i + 2, column=10).value = job_datum[9]
    ws.cell(row=i + 2, column=11).value = job_datum[10]
    ws.cell(row=i + 2, column=12).value = job_datum[11]
    ws.cell(row=i + 2, column=13).value = job_datum[12]
    ws.cell(row=i + 2, column=14).value = job_datum[13]
    ws.cell(row=i + 2, column=15).value = job_datum[14]
    ws.cell(row=i + 2, column=16).value = job_datum[15]
    ws.cell(row=i + 2, column=17).value = job_datum[16]
    ws.cell(row=i + 2, column=18).value = job_datum[17]


  # Save the spreadsheet.
  wb.save(filename)

def main():
    """
    The main function.
    """

    version = get_version()
    print("The version of the Selenium library is:", version)

    email = "singeryaseen79@gmail.com"
    password = "Y4seen19"
    keywords = ["Python", "SQL", "C#","C++","Java","AWS","Azure","Tableau","Power BI","Excel","Bachelor's","Master's","PhD"]
    driver = webdriver.Chrome()

    login_to_linkedin(driver, email, password)

    all_job_data = []

    has_next_page = True
    
    search_for_jobs(driver)
    while has_next_page:
      job_ids = get_job_ids(driver)
      for id in job_ids:
        time.sleep(4)
        card_container = driver.find_element(By.CLASS_NAME, "scaffold-layout__list-container")
        card_container.find_element(By.ID, id).click()
        time.sleep(0.5)
        card_container.find_element(By.ID, id).click()
        
        job_description = driver.find_element(By.ID, "job-details").text
        job_info_card = driver.find_element(By.CLASS_NAME,"job-details-jobs-unified-top-card__content--two-pane")

        job_title_element = job_info_card.find_element(By.CSS_SELECTOR, "h2")
        job_title = job_title_element.text
        job_details_elements = driver.find_element(By.CLASS_NAME, "job-details-jobs-unified-top-card__primary-description")
        company_name, location, time_of_post, number_of_applicants = split_job_post( job_details_elements.text)

        keyword_presence = check_words(job_description, keywords)

        all_job_data.append([job_title,company_name,location,time_of_post,number_of_applicants,keyword_presence[0],keyword_presence[1],keyword_presence[2],keyword_presence[3],keyword_presence[4]
                            ,keyword_presence[5],keyword_presence[6],keyword_presence[7],keyword_presence[8],keyword_presence[9],keyword_presence[10],keyword_presence[11],keyword_presence[12]])

      has_next_page = go_to_next_page(driver)
    # save it to Excel.
    print(all_job_data)
    save_to_excel(all_job_data,"job_data.xlsx",keywords)


if __name__ == "__main__":
  main()